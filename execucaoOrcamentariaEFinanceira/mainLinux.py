import pandas as pd
from openpyxl import load_workbook
import streamlit as st

def process_998(file_998):
    df_998 = pd.read_excel(file_998, dtype={'cod_contabil': str})
    print("DataFrame 998:")
    print(df_998.head())  
    return df_998

def process_3500(file_3500):
    df_3500 = pd.read_excel(file_3500, dtype={'fonte': str})
    print("DataFrame 3500:")
    print(df_3500.head())  
    return df_3500

def get_value_from_998(df, cod_contabil, col):
    filtered_df = df[df['cod_contabil'] == cod_contabil]
    result = filtered_df[col].sum()
    print(f"Debug 998 - Cod: {cod_contabil}, Col: {col}, Filtered DataFrame: {filtered_df}, Value: {result}")
    return result if not pd.isna(result) else 0

def get_values_from_3500(df, fontes):
    fontes_list = [f.strip() for f in fontes.split(',')]
    print(f"Debug 3500 - Fontes list: {fontes_list}")
    filtered_df = df[df['fonte'].isin(fontes_list)]
    result = filtered_df['pag_ate_mes'].sum()
    print(f"Debug 3500 - Filtered DataFrame: {filtered_df}, Value: {result}")
    return result if not pd.isna(result) else 0

def update_fundos_municipais(file_fundos, df_998, df_3500, aba):
    wb = load_workbook(file_fundos)
    ws = wb[aba]

   
    mappings = [
        ('998', '621200000000000', 'credito_atual', 3),
        ('998', '622130400000000', 'credito_atual', 4),
        ('998', '621300000000000', 'debito_atual', 6),
        ('998', '111111900000000', 'debito_atual', 7),
        ('998', '111115000000000', 'debito_atual', 8),
        ('3500', '1500000, 1500701, 1500702, 2500000, 2500701, 2500702, 31500000, 31500701, 31500702, 32500000, 51500000, 81500811, 81500821, 91500000', 'pag_ate_mes', 10)
    ]

    for mapping in mappings:
        relatorio, identificador, col_valor, row_idx = mapping
        if relatorio == '998':
            value = get_value_from_998(df_998, identificador, col_valor)
        elif relatorio == '3500':
            value = get_values_from_3500(df_3500, identificador)
        else:
            value = 0

        print(f"Updating cell at row {row_idx}, column 3 with value {value}")
        ws.cell(row=row_idx, column=3, value=value)  
    
    
    new_file_fundos = "FUNDOS_MUNICIPAIS_PRENCHIDO.xlsx"
    wb.save(new_file_fundos)
    print(f"Dados salvos em: {new_file_fundos}")
    return new_file_fundos

def main():
    st.title("RELATÓRIO EXECUÇÃO ORÇAMENTÁRIA E FINANCEIRA")

    file_998 = st.file_uploader("Upload do arquivo 998", type="xlsx")
    file_3500 = st.file_uploader("Upload do arquivo 3500", type="xlsx")
    file_fundos = st.file_uploader("Upload do arquivo Fundos Municipais", type="xlsx")
    aba = st.selectbox("Selecione a aba para preencher", 
                       ["Saúde", "Educação", "Saneamento", "Idoso", "Criança e do Adolescente", "Meio Ambiente"])
    
    if st.button("Preencher Dados"):
        if file_998 and file_3500 and file_fundos:
            df_998 = process_998(file_998)
            df_3500 = process_3500(file_3500)
            new_file_fundos = update_fundos_municipais(file_fundos, df_998, df_3500, aba)
            st.success(f"Dados preenchidos na aba {aba} com sucesso!")
            st.download_button(label="Baixar Arquivo Preenchido", data=open(new_file_fundos, 'rb').read(), file_name="FUNDOS_MUNICIPAIS_PRENCHIDO.xlsx")
        else:
            st.error("Por favor, faça o upload de todos os arquivos.")

if __name__ == "__main__":
    main()
