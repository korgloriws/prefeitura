import streamlit as st
import pandas as pd
import re
import openpyxl
import tempfile
import os


def normalizar_num_str(valor: object, keep_x: bool = False) -> str:

    if pd.isna(valor) or valor == '':
        return ''


    txt = str(valor).strip()
    if txt.replace('.', '').isdigit():
        txt = txt.split('.')[0]


    pattern = r'[^0-9Xx]' if keep_x else r'\D'
    txt = re.sub(pattern, '', txt)
    return txt.upper()


def to_norm_list(celula: object, keep_x: bool = False) -> list[str]:

    if pd.isna(celula) or celula == '':
        return []
    itens = re.split(r'[,\;/\s]+', str(celula))
    return [normalizar_num_str(i, keep_x) for i in itens if i.strip()]



def construir_mascara_codigos(series_codigos: pd.Series,
                               padroes_planilha: list[str]) -> pd.Series:

    regexes = []
    for cod in padroes_planilha:
        if not cod:
            continue
 
        cod_regex = cod.replace('X', r'\d').replace('x', r'\d')

        regexes.append(rf'^{cod_regex}\d*$')

    if not regexes:
        return pd.Series(False, index=series_codigos.index)

    big_regex = '|'.join(regexes)
    return series_codigos.str.match(big_regex, na=False)



def soma_valores_correspondentes(df_csv: pd.DataFrame,
                                 codigos_excel: list[str],
                                 co_excel: object,
                                 fonte_excel: object,
                                 debug: bool = False) -> float:

    mascara_codigo = construir_mascara_codigos(
        df_csv['CODIGO_RECEITA_NORM'], codigos_excel
    )


    lista_co = to_norm_list(co_excel)
    mascara_co = df_csv['CO_NORM'].isin(lista_co) if lista_co else True


    lista_fonte = to_norm_list(fonte_excel)
    mascara_fonte = df_csv['FONTE_NORM'].isin(lista_fonte) if lista_fonte else True

    filtro_final = mascara_codigo & mascara_co & mascara_fonte
    soma = df_csv.loc[filtro_final, 'VR_ARREC_MES_FONTE'].sum()

    if debug:
        st.write(
            f"Debug → Códigos {codigos_excel} | "
            f"CO {lista_co or 'qualquer'} | "
            f"Fonte {lista_fonte or 'qualquer'} | "
            f"Linhas={filtro_final.sum()} | Soma={soma}"
        )

    return soma



def processar_arquivos(csv_file, excel_file, debug: bool = False) -> str:

    df_csv = pd.read_csv(
        csv_file,
        sep=';',
        encoding='cp1252',
        usecols=["CODIGO_RECEITA", "FONTE_RECURSO", "CO", "VR_ARREC_MES_FONTE"]
    )

    df_csv['VR_ARREC_MES_FONTE'] = (
        df_csv['VR_ARREC_MES_FONTE']
        .astype(str)
        .str.replace('.', '', regex=False)
        .str.replace(',', '.', regex=False)
        .astype(float)
    )

    df_csv['CODIGO_RECEITA_NORM'] = df_csv['CODIGO_RECEITA'].apply(
        lambda x: normalizar_num_str(x, keep_x=True)
    )
    df_csv['CO_NORM']     = df_csv['CO'].apply(normalizar_num_str)
    df_csv['FONTE_NORM']  = df_csv['FONTE_RECURSO'].apply(normalizar_num_str)


    excel_path = tempfile.mktemp(suffix='.xlsx')
    with open(excel_path, 'wb') as tmp:
        tmp.write(excel_file.getvalue())

    book  = openpyxl.load_workbook(excel_path)
    sheet = book['Rascunho (2)']  


    for row in range(8, 42):  
        codigo_excel        = sheet.cell(row=row, column=2).value  
        co_excel            = sheet.cell(row=row, column=3).value  
        fonte_recurso_excel = sheet.cell(row=row, column=4).value  

        if not codigo_excel:
            if debug:
                st.write(f"Linha {row}: sem código — ignorada.")
            continue

        codigos_excel_norm = to_norm_list(codigo_excel, keep_x=True)
        soma = soma_valores_correspondentes(
            df_csv,
            codigos_excel_norm,
            co_excel,
            fonte_recurso_excel,
            debug
        )


        sheet.cell(row=row, column=5).value = round(soma, 2)


    download_path = tempfile.mktemp(suffix='.xlsx')
    book.save(download_path)
    book.close()
    return download_path



def main():
    st.title("Receita Corrente Líquida")

    csv_file   = st.file_uploader("1) Carregar arquivo QGR (CSV)", type=['csv'])
    excel_file = st.file_uploader("2) Carregar modelo Receita Corrente Líquida (Excel)", type=['xlsx', 'xls'])
    debug      = st.checkbox("Mostrar detalhes do processamento (debug)")

    if st.button("Processar Arquivos"):
        if not (csv_file and excel_file):
            st.error("Por favor, carregue ambos os arquivos antes de prosseguir.")
            return

        download_path = processar_arquivos(csv_file, excel_file, debug)

        with open(download_path, 'rb') as f:
            st.download_button(
                label="Baixar planilha preenchida",
                data=f,
                file_name="ReceitaCorrenteLiquida.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        os.remove(download_path)


if __name__ == '__main__':
    main()
