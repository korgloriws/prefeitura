from openpyxl import load_workbook
import os
import shutil
import tempfile
import zipfile
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter 

def brazilian_formatter(x):
    s = f"{int(x * 100):,}".replace(",", ".")  
    if len(s) <= 2:
        return f"0,{s.zfill(2)}"
    else:
        decimal_part = s[-2:]
        int_part = s[:-2]
        return f"{int_part},{decimal_part}"


def save_uploaded_file(uploaded_file):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            shutil.copyfileobj(uploaded_file, tmp_file)
            return tmp_file.name
    except Exception as e:
        st.error(f"Erro ao salvar o arquivo carregado: {e}")
        return None

def set_value(ws, address, value):
   
    cell = ws[address]
    for range_ in ws.merged_cells.ranges:
        if address in range_:
            
            ws.unmerge_cells(str(range_))
            
            top_left_cell_address = get_column_letter(range_.min_col) + str(range_.min_row)
            top_left_cell = ws[top_left_cell_address]
            top_left_cell.value = value
            ws.merge_cells(str(range_))
            return
    
    cell.value = value
  

def fill_and_save_form(group, name, payment_month, payment_date, template_file_path):
    try:
        output_file_name = f'Preenchido_{name}_Anexo_V - Comprovante de Retenção.xlsx'
        output_file_path = os.path.join(tempfile.gettempdir(), output_file_name)
        shutil.copy(template_file_path, output_file_path)

        wb = load_workbook(output_file_path)
        ws = wb.active  

        set_value(ws, 'A9', group['cpf_cnpj'].iloc[0])
        set_value(ws, 'D9', name)
        set_value(ws, 'F11', payment_month)
        set_value(ws, 'D40', payment_date)
        set_value(ws, 'A40', "Hemerson Fernandes Soares")

        start_row = 12
        group_sum = group.groupby('Código de receita').sum().reset_index()

        for i, (_, row) in enumerate(group_sum.iterrows()):
            row_number = start_row + i
            set_value(ws, f'C{row_number}', row['Código de receita'])
            set_value(ws, f'E{row_number}', brazilian_formatter(row['valor_bruto']))
            set_value(ws, f'G{row_number}', brazilian_formatter(row['valor_des']))
            set_value(ws, f'A{row_number}', payment_month)

        wb.save(output_file_path)
        wb.close()
        return output_file_path
    except Exception as e:
        st.error(f"Failed to process group {name}. Error: {e}")
        return None




def main():
    st.title("Processador de Formulários de Pagamento/IR RFB")

    uploaded_file = st.file_uploader("Escolha um arquivo Excel com os dados", type="xls")
    template_file = st.file_uploader("Escolha o arquivo de modelo do formulário", type="xlsx")
    payment_month = st.text_input("Insira o mês de pagamento:", "")
    payment_date = st.text_input("Insira a data do pagamento (dd/mm/aaaa):", "")
    file_name = st.text_input("Nome do arquivo para download:", "Preenchido_Formulario")

    process_button = st.button("Processar")

    if process_button and uploaded_file and template_file and payment_month and payment_date:
        template_file_path = save_uploaded_file(template_file)
        if template_file_path:
            source_df = pd.read_excel(uploaded_file)
            filtered_source_df = source_df[["credor", "cpf_cnpj", "valor_bruto", "valor_des", "Código de receita"]]
            grouped = filtered_source_df.groupby("credor")

            output_files = []
            for i, (name, group) in enumerate(grouped):
                st.write(f"Processando grupo {i+1}/{len(grouped)}: {name}")
                group = group.reset_index(drop=True)
                output_file = fill_and_save_form(group, name, payment_month, payment_date, template_file_path)
                if output_file:
                    output_files.append(output_file)

            if output_files:
                zip_file_path = os.path.join(tempfile.gettempdir(), f"{file_name}.zip")
                with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                    for file in output_files:
                        zipf.write(file, os.path.basename(file))
                        os.remove(file)  

                with open(zip_file_path, 'rb') as f:
                    st.download_button("Baixar Arquivos", f, file_name=f"{file_name}.zip")

if __name__ == "__main__":
    main()
