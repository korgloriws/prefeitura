import streamlit as st
import pandas as pd
import openpyxl
import re
import io

def regex_for_wildcard(text):
    if 'XX' in text:
        return '^' + text.replace('XX', '(\\d{2}|\\d)') + '$'
    else:
        return '^' + text + '$'

def adjust_formula(formula, row_offset):
    cell_references = re.findall(r'([A-Z]+)(\d+)', formula)
    for col, row in cell_references:
        new_row = int(row) + row_offset
        formula = formula.replace(f"{col}{row}", f"{col}{new_row}")
    return formula

def main():
    st.title("Conferência Despesa de Pessoal")
    
    file_data = st.file_uploader("Selecione o arquivo de dados:", type=["xlsx"])
    file_layout = st.file_uploader("Selecione o arquivo de layout:", type=["xlsx"])
    
    if file_data and file_layout:
        data_df = pd.read_excel(file_data)
        layout_df = pd.read_excel(file_layout, header=None)
        
        header_index = layout_df[layout_df.iloc[:, 2] == 'despesa'].index[0]
        layout_df.columns = layout_df.iloc[header_index]
        layout_df = layout_df.iloc[header_index+1:]  

        wb_original = openpyxl.load_workbook(file_layout, data_only=False)
        ws_original = wb_original.active
        
        for index, row in layout_df.iterrows():
            value = 0
            
            despesa_regex_pattern = regex_for_wildcard(str(row['despesa']))
            item_despesa_regex_pattern = regex_for_wildcard(str(row['item_despesa']))
            fontes = [f.strip() for f in str(row['fonte']).split(',') if f.strip()] if pd.notna(row['fonte']) else ['Todas fontes']

            for _, grouped_row in data_df.iterrows():
                despesa_match = re.match(despesa_regex_pattern, str(grouped_row['despesa']))
                item_despesa_match = re.match(item_despesa_regex_pattern, str(grouped_row['item_despesa']))

                if despesa_match and item_despesa_match:
                    if 'Todas fontes' in fontes or str(grouped_row['fonte']) in fontes:
                        value += grouped_row['vr_liq_mes']
            
            col = 6  
            ws_original.cell(row=index + 2 - header_index, column=col, value=value)
        
        ws_summary = wb_original.create_sheet("Somatórios")
        filtered_data = data_df[data_df['despesa'].astype(str).str.startswith('31')]
        total_sum = filtered_data['vr_liq_mes'].sum()
        ws_summary.append(["Despesa Início 31", "Total"])
        ws_summary.append(["31", total_sum])
        
        output = io.BytesIO()
        wb_original.save(output)
        output.seek(0)

        final_file_name = "Final_Conferência_Despesa_de_Pessoal_Layout.xlsx"

        st.download_button(
            label="Baixar arquivo final",
            data=output,
            file_name=final_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        return layout_df

if __name__ == '__main__':
    df = main()
