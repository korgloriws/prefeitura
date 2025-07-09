import streamlit as st
import pandas as pd
import warnings
import shutil
import os
from openpyxl import load_workbook


warnings.filterwarnings('ignore')


def processar_arquivos(script_xlsx, posicao_csv):
    df_script = pd.read_excel(script_xlsx, sheet_name='Em ordem', engine='openpyxl')
    df_posicao = pd.read_csv(posicao_csv, encoding='ISO-8859-1', sep=';', dtype=str)
    
    if len(df_script.columns) < 7:
        st.error(f"O arquivo Script precisa ter pelo menos 7 colunas. Atual: {len(df_script.columns)}")
        return None

    while len(df_posicao.columns) < 15:
        df_posicao[f'Coluna_{len(df_posicao.columns)}'] = ''

    col_script_id = df_script.iloc[:, 6]  
    col_script_d = df_script.iloc[:, 3]   
    col_script_e = df_script.iloc[:, 4]  

    for idx_script, id_script in col_script_id.items():
        matches = df_posicao.index[df_posicao.iloc[:, 3] == id_script].tolist()
        for idx_pos in matches:
            df_posicao.iat[idx_pos, 13] = col_script_d[idx_script]
            df_posicao.iat[idx_pos, 14] = col_script_e[idx_script]

    return df_posicao


def converter_valor_brasileiro(valor_str):
    if pd.isna(valor_str) or valor_str == '':
        return 0.0
    
    valor_str = str(valor_str).strip()
    valor_str = valor_str.replace('.', '').replace(',', '.')
    
    try:
        return float(valor_str)
    except:
        return 0.0


def processar_arquivos_adicionais(workbook, restos_nao_processados=None, restos_processados=None, ded_938=None):
    if not any([restos_nao_processados, restos_processados, ded_938]):
        return workbook
    
    worksheet_script = workbook['Script']
    
    if restos_nao_processados:
        try:
            df_restos_nao = pd.read_excel(restos_nao_processados, engine='openpyxl')
            if len(df_restos_nao.columns) >= 19:
                ultimo_valor = df_restos_nao.iloc[-1, 18]
                valor_final = float(ultimo_valor) if pd.notna(ultimo_valor) else 0.0
                
                for row_idx in range(1, worksheet_script.max_row + 1):
                    cell_value = worksheet_script.cell(row=row_idx, column=1).value
                    if cell_value and "Restos a pagar não Processados" in str(cell_value):
                        worksheet_script.cell(row=row_idx, column=2, value=valor_final)
                        break
        except Exception as e:
            st.warning(f"Erro ao processar Restos a Pagar não Processados: {e}")
    
    if restos_processados:
        try:
            df_restos_proc = pd.read_excel(restos_processados, engine='openpyxl')
            if len(df_restos_proc.columns) >= 19:
                ultimo_valor = df_restos_proc.iloc[-1, 18]
                valor_final = float(ultimo_valor) if pd.notna(ultimo_valor) else 0.0
                
                for row_idx in range(1, worksheet_script.max_row + 1):
                    cell_value = worksheet_script.cell(row=row_idx, column=1).value
                    if cell_value and "Restos a pagar Processados" in str(cell_value):
                        worksheet_script.cell(row=row_idx, column=2, value=valor_final)
                        break
        except Exception as e:
            st.warning(f"Erro ao processar Restos a Pagar Processados: {e}")
    
    if ded_938:
        try:
            df_ded = pd.read_excel(ded_938, engine='openpyxl')
            
            if len(df_ded.columns) >= 24:
                ultimo_valor_empenhado = df_ded.iloc[-1, 23]
                valor_empenhado_final = float(ultimo_valor_empenhado) if pd.notna(ultimo_valor_empenhado) else 0.0
                
                for row_idx in range(1, worksheet_script.max_row + 1):
                    cell_value = worksheet_script.cell(row=row_idx, column=1).value
                    if cell_value and "Valor Empenhado a Liquidar" in str(cell_value):
                        worksheet_script.cell(row=row_idx, column=2, value=valor_empenhado_final)
                        break
            
            if len(df_ded.columns) >= 25:
                ultimo_valor_liquidado = df_ded.iloc[-1, 24]
                valor_liquidado_final = float(ultimo_valor_liquidado) if pd.notna(ultimo_valor_liquidado) else 0.0
                
                for row_idx in range(1, worksheet_script.max_row + 1):
                    cell_value = worksheet_script.cell(row=row_idx, column=1).value
                    if cell_value and "Valor Liquidado a Pagar" in str(cell_value):
                        worksheet_script.cell(row=row_idx, column=2, value=valor_liquidado_final)
                        break
        except Exception as e:
            st.warning(f"Erro ao processar DED_938: {e}")
    
    return workbook


def processar_segunda_parte(df_posicao_atualizado, script_xlsx):
    col_g_valores = df_posicao_atualizado.iloc[:, 6]     
    col_n_funcoes = df_posicao_atualizado.iloc[:, 13]     
    col_o_subfuncoes = df_posicao_atualizado.iloc[:, 14]  
    
    dados_agrupados = []
    
    for idx in range(len(df_posicao_atualizado)):
        valor = col_g_valores.iloc[idx]
        funcao = col_n_funcoes.iloc[idx]
        subfuncao = col_o_subfuncoes.iloc[idx]
        
        if pd.isna(subfuncao) or subfuncao == '' or str(subfuncao).strip() == '':
            descricao = funcao
            if pd.notna(descricao) and str(descricao).strip() != '' and pd.notna(valor) and str(valor).strip() != '':
                dados_agrupados.append({
                    'descricao': descricao,
                    'valor': valor
                })
        else:
            funcao_str = str(funcao).strip() if pd.notna(funcao) else ''
            subfuncao_str = str(subfuncao).strip() if pd.notna(subfuncao) else ''
            
            if funcao_str == subfuncao_str and funcao_str != '':
                if pd.notna(funcao) and str(funcao).strip() != '' and pd.notna(valor) and str(valor).strip() != '':
                    dados_agrupados.append({
                        'descricao': funcao,
                        'valor': valor
                    })
            else:
                if pd.notna(funcao) and str(funcao).strip() != '' and pd.notna(valor) and str(valor).strip() != '':
                    dados_agrupados.append({
                        'descricao': funcao,
                        'valor': valor
                    })
                
                if pd.notna(subfuncao) and str(subfuncao).strip() != '' and pd.notna(valor) and str(valor).strip() != '':
                    dados_agrupados.append({
                        'descricao': subfuncao,
                        'valor': valor
                    })
    
    df_agrupado = pd.DataFrame(dados_agrupados)
    if len(df_agrupado) > 0:
        df_agrupado['valor'] = df_agrupado['valor'].apply(converter_valor_brasileiro)
        df_soma = df_agrupado.groupby('descricao')['valor'].sum().reset_index()
        
        workbook = load_workbook(script_xlsx)
        worksheet_script = workbook['Script']
        
        for _, row in df_soma.iterrows():
            descricao = row['descricao']
            valor_soma = row['valor']
            
            for row_idx in range(1, worksheet_script.max_row + 1):
                cell_value = worksheet_script.cell(row=row_idx, column=1).value
                if cell_value and str(cell_value).strip() == str(descricao).strip():
                    worksheet_script.cell(row=row_idx, column=2, value=valor_soma)
                    break
        
        return workbook
    
    return None


def main():
    st.title('Disponibilidade de Caixa - Sistema Completo')
    st.write('Faça upload dos arquivos:')
    script_xlsx = st.file_uploader('Script da Disponibilidade de Caixa (.xlsx)', type=['xlsx'])
    posicao_csv = st.file_uploader('Posição de Bancos (.csv)', type=['csv'])
    
    st.write("### Arquivos Adicionais (Opcionais):")
    restos_nao_processados = st.file_uploader('Restos a Pagar não Processados (.xlsx)', type=['xlsx'])
    restos_processados = st.file_uploader('Restos a Pagar Processados (.xlsx)', type=['xlsx'])
    ded_938 = st.file_uploader('DED_938 (.xlsx)', type=['xlsx'])

    if script_xlsx and posicao_csv:
        resultado = processar_arquivos(script_xlsx, posicao_csv)
        
        if resultado is not None:
            df_posicao_atualizado = resultado
            st.success('Primeira parte concluída!')
            
            workbook_final = processar_segunda_parte(df_posicao_atualizado, script_xlsx)
            
            if workbook_final is not None:
                workbook_final = processar_arquivos_adicionais(
                    workbook_final, 
                    restos_nao_processados, 
                    restos_processados, 
                    ded_938
                )
                
                st.success('Processamento completo concluído!')
                
                col1, col2 = st.columns(2)
                
                with col1:
                    csv_posicao = df_posicao_atualizado.to_csv(index=False, encoding='ISO-8859-1', sep=';').encode('ISO-8859-1')
                    st.download_button('Baixar posição de bancos atualizado', csv_posicao, 'posicaoDeBancosAtualizado.csv', 'text/csv')
                
                with col2:
                    from io import BytesIO
                    buffer = BytesIO()
                    workbook_final.save(buffer)
                    buffer.seek(0)
                    st.download_button('Baixar arquivo Excel original atualizado', buffer, 'Script_Disponibilidade_Caixa_Atualizado.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    main() 